import urllib.request
from bs4 import BeautifulSoup
import requests
import pandas as pd
from datetime import datetime
import pandas
from openpyxl import load_workbook

URLS_FED_PROJ_1 = ['https://spending.gov.ru/np/D/fp/D1/grbs/',]

URLS_FED_PROJ_2 = ['https://spending.gov.ru/np/D/fp/D2/grbs/',]

URLS_FED_PROJ_3 = ['https://spending.gov.ru/np/D/fp/D3/grbs/',]

URLS_FED_PROJ_4 = ['https://spending.gov.ru/np/D/fp/D4/grbs/',]

URLS_FED_PROJ_5 = ['https://spending.gov.ru/np/D/fp/D5/grbs/',]

URLS_FED_PROJ_6 = ['https://spending.gov.ru/np/D/fp/D6/grbs/',]


# закрепить их за программами
proj_dict = {'Нормативное регулирование цифровой среды': URLS_FED_PROJ_1,
            'Информационная инфраструктура': URLS_FED_PROJ_2,
            'Кадры для цифровой экономики': URLS_FED_PROJ_3,
            'Информационная безопасность': URLS_FED_PROJ_4,
            'Цифровые технологии': URLS_FED_PROJ_5,
            'Цифровое государственное управление': URLS_FED_PROJ_6}

# вспомогательные функции
def clean_text(my_str: str) -> float:
    my_str = my_str.strip()
    my_str = my_str.replace(',', '.')
    my_str = my_str.replace('₽', '')
    my_str = my_str.replace(u'\xa0', '')
    return(float(my_str))

def clean_share(my_str: str) -> float:
    my_str = my_str.strip()
    my_str = my_str.replace(',', '.')
    my_str = my_str.replace('%', '')
    my_str = my_str.replace(u'\xa0', '')
    return(float(my_str))


def str_to_date(str_date) -> datetime.date:
    str_date = str_date.split()[:3]
    month_list = {'ноября': '11',
             'декабря': '12',
              'января': '01',
             'февраля': '02',
             'марта': '03',
             'апреля': '04',
             'мая': '05',
             'июня': '06',
             'июля': '07',
             'августа': '08',
             'сентября': '09',
             'октября': '10'}
    str_date = list([int(str_date[2]),
            int(month_list.get(str_date[1])),
            int(str_date[0]),])
                 
    str_date = datetime(str_date[0], str_date[1], str_date[2])
    return(datetime.date(str_date))


# основная функция с правкой
def parsing_func(URL:str) -> pd.DataFrame():
    my_df = pd.DataFrame(columns=['ГРБС',
                                 'объем_средств',
                                 'доля',])
    html_content = requests.get(URL).text
    bsObj = BeautifulSoup(html_content, 'html.parser')
    table_body = bsObj.find('tbody')

    data = []


    rows = table_body.find_all('tr')
    for row in rows:
        cols = row.find_all('td')
        cols = [ele.text.strip() for ele in cols]
        data.append([ele for ele in cols if ele])
        for rec in data:
            my_dict = {'ГРБС': rec[0],
                       'объем_средств': clean_text(rec[1]),
                       'доля': clean_share(rec[2]),}
        
        my_df = my_df.append(my_dict, ignore_index=True)
    return(my_df)

sheet = 'грбс'

    
def parse_n_save(key: str, sheet:str=sheet) -> None:
    URLS = proj_dict.get(key)
    name = key+'.xlsx'
    print('Working on {}'.format(key))
    if len(URLS) > 0:
        main_df = pd.DataFrame(columns=['ГРБС',
                                 'объем_средств',
                                 'доля',])
        
        for URL in URLS:
            temp_df = parsing_func(URL)
            main_df = pd.concat([main_df, temp_df])
        
        print(f"{len(main_df)} records added")
        main_df.reset_index(inplace=True)
        with pandas.ExcelWriter(name, engine='openpyxl') as writer:
            writer.book = load_workbook(name)
            main_df.to_excel(writer, sheet_name=sheet)
        writer.save()
        
        
    else:
        main_df = parsing_func(URLS[0])
        print(f"{len(main_df)} records added")    

        with pandas.ExcelWriter(name, engine='openpyxl') as writer:
            writer.book = load_workbook(name)
            main_df.to_excel(writer, sheet_name=sheet)
        writer.save()
        
for key in proj_dict.keys():
    parse_n_save(key)